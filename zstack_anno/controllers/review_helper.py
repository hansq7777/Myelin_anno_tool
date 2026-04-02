from __future__ import annotations

from datetime import datetime
import csv
import os
import random
import re
import shutil
from typing import TYPE_CHECKING

import tifffile
from PyQt5.QtWidgets import QFileDialog, QMessageBox

from ..utils.review_utils import (
    build_pair_key,
    is_review_completed,
    normalize_review_grade,
    windows_to_local_path,
    local_to_windows_path,
)

if TYPE_CHECKING:  # pragma: no cover
    from .main_controller import MainController

DEFAULT_REVIEW_TRACKER_WINDOWS = (
    r"D:\Research\Image Analysis\Confocal Myelin data\zstack_annotation_tracker_2026-02-06.xlsx"
)

TRACKER_BASE_COLUMNS = [
    "zstack_id",
    "raw_path",
    "inference_path",
    "source_group",
    "pair_key",
    "pair_status",
]

TRACKER_REVIEW_COLUMNS = [
    "review_grade",
    "review_status",
    "review_note",
    "review_updated_at",
    "review_corrected_mask_path",
    "review_corrected_saved_at",
    "review_final_mask_path",
    "review_final_mask_source",
    "review_final_exported_at",
    "review_completed",
    "review_completed_at",
]

TRACKER_COLUMN_ORDER = TRACKER_BASE_COLUMNS + TRACKER_REVIEW_COLUMNS


class ReviewMixin:
    """Batch review helpers for raw+prediction quality control."""

    def _init_review_state(self: "MainController") -> None:
        self._review_tracker_path: str | None = None
        self._review_tracker_format: str = ""
        self._review_header_names: list[str] = []
        self._review_headers: dict[str, int] = {}
        self._review_table_rows: list[dict[str, object]] = []
        self._review_items: list[dict] = []
        self._review_filtered_indices: list[int] = []
        self._review_current_item_index: int = -1
        self._review_current_filtered_pos: int = -1
        self._review_grade_updating: bool = False
        self._review_filter_updating: bool = False
        self._review_skipped_missing_paths: int = 0
        self._review_skipped_missing_files: int = 0
        self._review_random_enabled: bool = True
        self._review_random_seen_indices: set[int] = set()
        self._review_random_history: list[int] = []
        self._review_rng = random.Random()

    # --------- public actions ---------
    def _open_review_tracker(self: "MainController") -> None:
        default_local = windows_to_local_path(DEFAULT_REVIEW_TRACKER_WINDOWS)

        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open Review Tracker",
            self._review_tracker_path or default_local or "",
            "Tracker Files (*.xlsx *.csv);;Excel Files (*.xlsx);;CSV Files (*.csv)",
        )
        if not path:
            return
        self._load_review_tracker(path)

    def _review_build_tracker_from_folders(self: "MainController") -> None:
        raw_dir = QFileDialog.getExistingDirectory(self, "Select Raw Stack Folder")
        if not raw_dir:
            return
        pred_dir = QFileDialog.getExistingDirectory(self, "Select Prediction Mask Folder")
        if not pred_dir:
            return

        default_tracker_path = self._review_default_tracker_path(raw_dir)
        tracker_path, _ = QFileDialog.getSaveFileName(
            self,
            "Create or Refresh Review Tracker",
            self._review_tracker_path or default_tracker_path,
            "Tracker Files (*.xlsx *.csv);;Excel Files (*.xlsx);;CSV Files (*.csv)",
        )
        if not tracker_path:
            return
        tracker_path = tracker_path.strip()
        if not tracker_path:
            return
        if not os.path.splitext(tracker_path)[1]:
            tracker_path += ".xlsx"

        try:
            rows, stats = self._review_build_tracker_rows(raw_dir, pred_dir, tracker_path)
            self._review_write_tracker(tracker_path, TRACKER_COLUMN_ORDER, rows)
        except Exception as exc:
            QMessageBox.warning(self, "Build Tracker", f"Failed to build tracker:\n{exc}")
            return

        self._load_review_tracker(tracker_path)
        QMessageBox.information(
            self,
            "Build Tracker",
            (
                f"Tracker updated: {tracker_path}\n"
                f"Matched: {stats['matched']}\n"
                f"Raw only: {stats['raw_only']}\n"
                f"Prediction only: {stats['pred_only']}\n"
                f"Raw duplicates: {stats['raw_duplicates']}\n"
                f"Prediction duplicates: {stats['pred_duplicates']}"
            ),
        )

    def _review_prev_stack(self: "MainController") -> None:
        if not self._review_filtered_indices:
            return
        if self._review_should_random_select():
            if len(self._review_random_history) < 2:
                return
            self._review_random_history.pop()
            prev_item_index = self._review_random_history[-1]
            if prev_item_index not in self._review_filtered_indices:
                return
            pos = self._review_filtered_indices.index(prev_item_index)
            self._review_load_filtered_pos(pos, force=True, record_history=False)
            return
        self._review_load_filtered_pos(self._review_current_filtered_pos - 1)

    def _review_next_stack(self: "MainController") -> None:
        if not self._review_filtered_indices:
            return
        if self._review_should_random_select():
            next_item_index = self._review_pick_random_item_index()
            if next_item_index is None:
                QMessageBox.information(
                    self,
                    "Review Queue",
                    "No unfinished items available for random selection.",
                )
                return
            pos = self._review_filtered_indices.index(next_item_index)
            self._review_load_filtered_pos(pos, force=True)
            return
        self._review_load_filtered_pos(self._review_current_filtered_pos + 1)

    def _review_mark_a(self: "MainController") -> None:
        self._review_set_grade("A", move_next=True)

    def _review_mark_b(self: "MainController") -> None:
        self._review_set_grade("B", move_next=True)

    def _review_mark_c(self: "MainController") -> None:
        self._review_set_grade("C", move_next=True)

    def _review_on_grade_combo_changed(self: "MainController", text: str) -> None:
        if self._review_grade_updating:
            return
        grade = normalize_review_grade(text)
        self._review_set_grade(grade, move_next=False)

    def _review_on_filter_changed(self: "MainController", _text: str) -> None:
        if self._review_filter_updating:
            return
        self._review_reset_random_state()
        self._review_rebuild_filtered_indices(load_item=True)

    def _review_save_corrected_mask(self: "MainController") -> None:
        if self.model.masks is None:
            QMessageBox.warning(self, "Save Corrected Mask", "No mask data to save.")
            return
        item = self._review_current_item()
        if item is None:
            return

        tracker_dir = os.path.dirname(self._review_tracker_path or "")
        grade = normalize_review_grade(item.get("grade"))
        grade_folder = grade if grade else "UNREVIEWED"
        out_dir = os.path.join(tracker_dir, "review_corrected_masks", grade_folder)
        os.makedirs(out_dir, exist_ok=True)
        zstack_id = item.get("zstack_id") or f"row{item['row']}"
        out_path = os.path.join(out_dir, f"{zstack_id}_review_mask.tif")

        if os.path.exists(out_path):
            ret = QMessageBox.question(
                self,
                "Save Corrected Mask",
                f"Overwrite existing file?\n{out_path}",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if ret != QMessageBox.Yes:
                return

        metadata = self._review_build_corrected_mask_metadata(item, out_path)
        self.model.save_masks(out_path, metadata=metadata)
        self._review_set_cell(item["row"], "review_corrected_mask_path", local_to_windows_path(out_path))
        self._review_set_cell(item["row"], "review_corrected_saved_at", self._review_now())
        self._review_mark_completed(item)
        item["corrected_path"] = out_path
        if self._review_save_tracker():
            self.statusBar().showMessage(
                f"Corrected mask saved and marked completed: {os.path.basename(out_path)}"
            )
        self._review_rebuild_filtered_indices(load_item=False)
        self._update_file_labels()

    def _review_export_final_masks(self: "MainController") -> None:
        if not self._review_items or not self._review_tracker_path:
            QMessageBox.warning(self, "Export Final Masks", "Please open a review tracker first.")
            return

        reviewed = [
            item for item in self._review_items
            if normalize_review_grade(item.get("grade")) in {"A", "B", "C"}
        ]
        if not reviewed:
            QMessageBox.information(
                self,
                "Export Final Masks",
                "No reviewed A/B/C items found. Mark stacks first, then export.",
            )
            return

        tracker_dir = os.path.dirname(self._review_tracker_path)
        export_root = os.path.join(tracker_dir, "review_final_masks")
        os.makedirs(export_root, exist_ok=True)

        expected_files: set[str] = set()
        copied = 0
        missing = 0

        for item in reviewed:
            grade = normalize_review_grade(item.get("grade"))
            if not grade:
                continue
            zstack_id = item.get("zstack_id") or f"row{item['row']}"
            corrected_path = item.get("corrected_path") or ""
            pred_path = item.get("pred_path") or ""

            if corrected_path and os.path.exists(corrected_path):
                src_path = corrected_path
                src_kind = "corrected"
            else:
                src_path = pred_path
                src_kind = "inference"

            if not src_path or not os.path.exists(src_path):
                missing += 1
                continue

            out_dir = os.path.join(export_root, grade)
            os.makedirs(out_dir, exist_ok=True)
            out_path = os.path.join(out_dir, f"{zstack_id}_final_mask.tif")
            expected_files.add(os.path.abspath(out_path))

            try:
                if os.path.abspath(src_path) != os.path.abspath(out_path):
                    shutil.copy2(src_path, out_path)
                copied += 1
            except Exception as exc:
                QMessageBox.warning(
                    self,
                    "Export Final Masks",
                    f"Failed to export mask for {zstack_id}:\n{exc}",
                )
                return

            self._review_set_cell(item["row"], "review_final_mask_path", local_to_windows_path(out_path))
            self._review_set_cell(item["row"], "review_final_mask_source", src_kind)
            self._review_set_cell(item["row"], "review_final_exported_at", self._review_now())

        stale_files: list[str] = []
        for root, _dirs, files in os.walk(export_root):
            for name in files:
                if not name.lower().endswith((".tif", ".tiff")):
                    continue
                path = os.path.abspath(os.path.join(root, name))
                if path not in expected_files:
                    stale_files.append(path)

        deleted = 0
        if stale_files:
            ret = QMessageBox.question(
                self,
                "Export Final Masks",
                (
                    f"Found {len(stale_files)} stale files in review_final_masks.\n"
                    "Delete them now?"
                ),
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if ret == QMessageBox.Yes:
                for path in stale_files:
                    try:
                        os.remove(path)
                        deleted += 1
                    except Exception:
                        continue

                for grade in ("A", "B", "C", "UNREVIEWED"):
                    grade_dir = os.path.join(export_root, grade)
                    if os.path.isdir(grade_dir) and not os.listdir(grade_dir):
                        try:
                            os.rmdir(grade_dir)
                        except Exception:
                            pass

        if self._review_save_tracker():
            self.statusBar().showMessage(
                f"Exported final masks: {copied} | missing: {missing} | deleted stale: {deleted}"
            )

    # --------- tracker loading ---------
    def _load_review_tracker(self: "MainController", path: str) -> None:
        path = (path or "").strip()
        if not path:
            return
        if not os.path.exists(path):
            QMessageBox.warning(self, "Review Tracker", f"Tracker not found:\n{path}")
            return

        try:
            header_names, table_rows, fmt = self._review_read_tracker(path)
        except Exception as exc:
            QMessageBox.warning(self, "Review Tracker", f"Failed to open tracker:\n{exc}")
            return

        if "zstack_id" not in header_names:
            QMessageBox.warning(self, "Review Tracker", "Missing required column: zstack_id")
            return

        self._review_tracker_path = path
        self._review_tracker_format = fmt
        self._review_header_names = list(header_names)
        self._review_table_rows = table_rows
        self._review_rebuild_header_index()
        self._review_normalize_table_rows()
        self._review_ensure_columns()
        self._review_reset_random_state()
        self._review_refresh_items()
        self._set_review_controls_enabled(bool(self._review_items))
        if not self._review_items:
            self._review_rebuild_filtered_indices(load_item=False)
            QMessageBox.information(
                self,
                "Review Tracker",
                (
                    "Tracker loaded, but 0 usable raw+prediction pairs were found.\n\n"
                    "Please check that path columns are populated and files exist.\n"
                    "Supported raw columns: raw_path, raw_found_path\n"
                    "Supported prediction columns: inference_found_path, inference_path, prediction_path, pred_path, mask_path"
                ),
            )
            self.statusBar().showMessage(
                (
                    "Review tracker loaded with 0 usable items "
                    f"(missing paths: {self._review_skipped_missing_paths}, "
                    f"missing files: {self._review_skipped_missing_files})."
                )
            )
            return

        default_filter = self._review_pick_initial_filter()
        self._review_set_filter(default_filter)
        self._review_rebuild_filtered_indices(load_item=True)
        self.statusBar().showMessage(
            (
                f"Review tracker loaded: {os.path.basename(path)} | "
                f"usable: {len(self._review_items)} | "
                f"missing paths: {self._review_skipped_missing_paths} | "
                f"missing files: {self._review_skipped_missing_files} | "
                f"start filter: {default_filter} | "
                f"selection: random-unfinished"
            )
        )

    def _review_read_tracker(self: "MainController", path: str) -> tuple[list[str], list[dict[str, object]], str]:
        ext = os.path.splitext(path)[1].lower()
        if ext not in {".xlsx", ".csv"}:
            raise ValueError("Unsupported tracker format. Use .xlsx or .csv")
        if ext == ".csv":
            header_names, rows = self._review_read_csv_tracker(path)
            return header_names, rows, "csv"
        header_names, rows = self._review_read_xlsx_tracker(path)
        return header_names, rows, "xlsx"

    def _review_read_xlsx_tracker(self: "MainController", path: str) -> tuple[list[str], list[dict[str, object]]]:
        try:
            import openpyxl  # pylint: disable=import-outside-toplevel
        except Exception as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(
                "openpyxl is required for .xlsx tracker support. Install with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header_names: list[str] = []
        col_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(1, col).value
            if isinstance(value, str) and value.strip():
                name = value.strip()
                header_names.append(name)
                col_map[name] = col

        rows: list[dict[str, object]] = []
        for row_idx in range(2, ws.max_row + 1):
            row: dict[str, object] = {}
            for name in header_names:
                val = ws.cell(row_idx, col_map[name]).value
                row[name] = "" if val is None else val
            rows.append(row)
        return header_names, rows

    def _review_read_csv_tracker(self: "MainController", path: str) -> tuple[list[str], list[dict[str, object]]]:
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            header_names = [h.strip() for h in (reader.fieldnames or []) if h and h.strip()]
            rows: list[dict[str, object]] = []
            for rec in reader:
                row = {name: (rec.get(name, "") if rec else "") for name in header_names}
                rows.append(row)
        return header_names, rows

    def _review_write_tracker(
        self: "MainController",
        path: str,
        header_names: list[str],
        rows: list[dict[str, object]],
    ) -> None:
        ext = os.path.splitext(path)[1].lower()
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        if ext == ".csv":
            with open(path, "w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=header_names, extrasaction="ignore")
                writer.writeheader()
                for row in rows:
                    writer.writerow({name: row.get(name, "") for name in header_names})
            return
        if ext != ".xlsx":
            raise ValueError("Unsupported tracker format. Use .xlsx or .csv")
        try:
            import openpyxl  # pylint: disable=import-outside-toplevel
        except Exception as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(
                "openpyxl is required for .xlsx tracker support. Install with: pip install openpyxl"
            ) from exc
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, name in enumerate(header_names, start=1):
            ws.cell(1, col).value = name
        for row_idx, row in enumerate(rows, start=2):
            for col, name in enumerate(header_names, start=1):
                value = row.get(name, "")
                ws.cell(row_idx, col).value = None if value == "" else value
        wb.save(path)

    def _review_refresh_items(self: "MainController") -> None:
        items: list[dict] = []
        skipped_missing_paths = 0
        skipped_missing_files = 0

        for row in self._review_row_numbers():
            zstack_id = self._review_cell(row, "zstack_id")
            if not zstack_id:
                continue
            raw_path = self._review_first_nonempty(
                row,
                ["raw_path", "raw_found_path"],
            )
            pred_path = self._review_first_nonempty(
                row,
                [
                    "inference_found_path",
                    "inference_path",
                    "prediction_path",
                    "pred_path",
                    "mask_path",
                ],
            )

            raw_local = windows_to_local_path(raw_path) if raw_path else ""
            pred_local = windows_to_local_path(pred_path) if pred_path else ""
            if not raw_local or not pred_local:
                skipped_missing_paths += 1
                continue
            if not os.path.exists(raw_local) or not os.path.exists(pred_local):
                skipped_missing_files += 1
                continue

            grade = normalize_review_grade(self._review_cell(row, "review_grade"))
            completed_at = self._review_cell(row, "review_completed_at")
            completed = is_review_completed(self._review_cell(row, "review_completed"), completed_at)
            corrected_path = windows_to_local_path(self._review_cell(row, "review_corrected_mask_path"))
            items.append(
                {
                    "row": row,
                    "zstack_id": zstack_id,
                    "raw_path": raw_local,
                    "pred_path": pred_local,
                    "grade": grade,
                    "completed": completed,
                    "completed_at": completed_at,
                    "corrected_path": corrected_path if corrected_path and os.path.exists(corrected_path) else "",
                    "source_group": self._review_cell(row, "source_group") or "",
                }
            )

        self._review_items = items
        self._review_skipped_missing_paths = skipped_missing_paths
        self._review_skipped_missing_files = skipped_missing_files

    def _review_ensure_columns(self: "MainController") -> None:
        added = False
        for name in TRACKER_REVIEW_COLUMNS:
            if name not in self._review_headers:
                self._review_header_names.append(name)
                self._review_rebuild_header_index()
                for row in self._review_table_rows:
                    row[name] = ""
                added = True
        if added:
            self._review_save_tracker()

    # --------- tracker build helpers ---------
    def _review_build_tracker_rows(
        self: "MainController", raw_dir: str, pred_dir: str, tracker_path: str
    ) -> tuple[list[dict[str, object]], dict[str, int]]:
        raw_files = self._review_collect_stack_files(raw_dir, allow_czi=True)
        pred_files = self._review_collect_stack_files(pred_dir, allow_czi=False)

        raw_map: dict[str, str] = {}
        pred_map: dict[str, str] = {}
        raw_duplicates = 0
        pred_duplicates = 0

        for path in raw_files:
            key = build_pair_key(os.path.basename(path), is_prediction=False)
            if not key:
                continue
            if key in raw_map:
                raw_duplicates += 1
                continue
            raw_map[key] = path

        for path in pred_files:
            key = build_pair_key(os.path.basename(path), is_prediction=True)
            if not key:
                continue
            if key in pred_map:
                pred_duplicates += 1
                continue
            pred_map[key] = path

        raw_keys = set(raw_map.keys())
        pred_keys = set(pred_map.keys())
        matched_keys = sorted(raw_keys & pred_keys)
        raw_only_keys = sorted(raw_keys - pred_keys)
        pred_only_keys = sorted(pred_keys - raw_keys)

        existing = self._review_load_existing_rows_by_id(tracker_path)
        rows: list[dict[str, object]] = []

        for key in matched_keys:
            row = self._review_new_tracker_row(
                zstack_id=key,
                raw_path=raw_map[key],
                pred_path=pred_map[key],
                source_group=self._review_guess_source_group(raw_dir, raw_map[key]),
                pair_status="matched",
            )
            self._review_merge_review_fields(row, existing.get(key))
            rows.append(row)

        for key in raw_only_keys:
            row = self._review_new_tracker_row(
                zstack_id=key,
                raw_path=raw_map[key],
                pred_path="",
                source_group=self._review_guess_source_group(raw_dir, raw_map[key]),
                pair_status="raw_only",
            )
            self._review_merge_review_fields(row, existing.get(key))
            rows.append(row)

        for key in pred_only_keys:
            row = self._review_new_tracker_row(
                zstack_id=key,
                raw_path="",
                pred_path=pred_map[key],
                source_group="",
                pair_status="pred_only",
            )
            self._review_merge_review_fields(row, existing.get(key))
            rows.append(row)

        stats = {
            "matched": len(matched_keys),
            "raw_only": len(raw_only_keys),
            "pred_only": len(pred_only_keys),
            "raw_duplicates": raw_duplicates,
            "pred_duplicates": pred_duplicates,
        }
        return rows, stats

    @staticmethod
    def _review_tracker_name_stem(raw_dir: str) -> str:
        normalized = os.path.normpath(raw_dir or "")
        raw_name = os.path.basename(normalized)
        parent_name = os.path.basename(os.path.dirname(normalized))
        generic_names = {
            "original_zstacks",
            "raw",
            "raw_stacks",
            "stacks",
            "images",
        }
        if raw_name.lower() in generic_names and parent_name:
            stem = f"{parent_name}_{raw_name}"
        else:
            stem = raw_name or parent_name or "zstacks"
        stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
        return stem or "zstacks"

    @classmethod
    def _review_default_tracker_filename(
        cls,
        raw_dir: str,
        *,
        when: datetime | None = None,
    ) -> str:
        created = (when or datetime.now()).strftime("%Y-%m-%d")
        stem = cls._review_tracker_name_stem(raw_dir)
        return f"{stem}_review_tracker_{created}.xlsx"

    @classmethod
    def _review_default_tracker_path(
        cls,
        raw_dir: str,
        *,
        when: datetime | None = None,
    ) -> str:
        return os.path.join(raw_dir, cls._review_default_tracker_filename(raw_dir, when=when))

    @staticmethod
    def _review_collect_stack_files(root_dir: str, *, allow_czi: bool) -> list[str]:
        out: list[str] = []
        for root, _dirs, files in os.walk(root_dir):
            for name in files:
                lower = name.lower()
                if lower.endswith((".tif", ".tiff")):
                    out.append(os.path.join(root, name))
                    continue
                if allow_czi and lower.endswith(".czi"):
                    out.append(os.path.join(root, name))
        out.sort()
        return out

    @staticmethod
    def _review_guess_source_group(root_dir: str, path: str) -> str:
        try:
            rel = os.path.relpath(path, root_dir)
        except Exception:
            return ""
        parts = rel.split(os.sep)
        return parts[0] if len(parts) > 1 else ""

    def _review_load_existing_rows_by_id(self: "MainController", tracker_path: str) -> dict[str, dict[str, object]]:
        if not os.path.exists(tracker_path):
            return {}
        try:
            header_names, rows, _ = self._review_read_tracker(tracker_path)
        except Exception:
            return {}
        if "zstack_id" not in header_names:
            return {}
        out: dict[str, dict[str, object]] = {}
        for row in rows:
            key = str(row.get("zstack_id", "")).strip().lower()
            if key and key not in out:
                out[key] = row
        return out

    @staticmethod
    def _review_new_tracker_row(
        *,
        zstack_id: str,
        raw_path: str,
        pred_path: str,
        source_group: str,
        pair_status: str,
    ) -> dict[str, object]:
        row = {name: "" for name in TRACKER_COLUMN_ORDER}
        row["zstack_id"] = zstack_id
        row["raw_path"] = local_to_windows_path(raw_path) if raw_path else ""
        row["inference_path"] = local_to_windows_path(pred_path) if pred_path else ""
        row["source_group"] = source_group
        row["pair_key"] = zstack_id
        row["pair_status"] = pair_status
        return row

    @staticmethod
    def _review_merge_review_fields(row: dict[str, object], old_row: dict[str, object] | None) -> None:
        if not old_row:
            return
        for name in TRACKER_REVIEW_COLUMNS:
            value = old_row.get(name, "")
            if value not in (None, ""):
                row[name] = value

    # --------- row actions ---------
    def _review_set_grade(self: "MainController", grade: str, move_next: bool) -> None:
        item = self._review_current_item()
        if item is None:
            return
        grade = normalize_review_grade(grade)
        row = item["row"]

        status_map = {
            "A": "accept_direct",
            "B": "accept_after_edit",
            "C": "reject_or_later",
            "": "unreviewed",
        }
        self._review_set_cell(row, "review_grade", grade or None)
        self._review_set_cell(row, "review_status", status_map.get(grade, "unreviewed"))
        self._review_set_cell(row, "review_updated_at", self._review_now())

        item["grade"] = grade
        self._review_save_tracker()
        self._review_update_info_label()

        if move_next:
            self._review_rebuild_filtered_indices(load_item=False)
            self._review_next_stack()
            return

        self._review_rebuild_filtered_indices(load_item=False)
        self._review_sync_grade_combo()
        self.statusBar().showMessage(f"Marked {item['zstack_id']} as {grade or 'UNREVIEWED'}")

    def _review_mark_completed(self: "MainController", item: dict) -> None:
        row = item["row"]
        now = self._review_now()
        self._review_set_cell(row, "review_completed", 1)
        self._review_set_cell(row, "review_completed_at", now)
        self._review_set_cell(row, "review_updated_at", now)
        item["completed"] = True
        item["completed_at"] = now

    # --------- navigation ---------
    def _review_rebuild_filtered_indices(self: "MainController", load_item: bool) -> None:
        mode = self.review_filter_combo.currentText() if hasattr(self, "review_filter_combo") else "All"
        previous_item_index = self._review_current_item_index
        previous_pos = self._review_current_filtered_pos

        filtered: list[int] = []
        for idx, item in enumerate(self._review_items):
            grade = normalize_review_grade(item.get("grade"))
            completed = bool(item.get("completed"))
            if mode == "All":
                filtered.append(idx)
            elif mode == "Unreviewed":
                if not grade and not completed:
                    filtered.append(idx)
            elif grade == mode:
                filtered.append(idx)

        self._review_filtered_indices = filtered
        if not filtered:
            self._review_current_item_index = -1
            self._review_current_filtered_pos = -1
            self._review_update_info_label()
            return

        if load_item and self._review_should_random_select():
            item_index = self._review_pick_random_item_index()
            if item_index is not None:
                pos = filtered.index(item_index)
                self._review_load_filtered_pos(pos, force=True)
                return

        if previous_item_index in filtered:
            pos = filtered.index(previous_item_index)
        else:
            pos = min(max(previous_pos, 0), len(filtered) - 1)

        self._review_current_filtered_pos = pos
        self._review_current_item_index = filtered[pos]
        if load_item:
            self._review_load_filtered_pos(pos, force=True)
        else:
            self._review_update_info_label()

    def _review_should_random_select(self: "MainController") -> bool:
        if not self._review_random_enabled:
            return False
        if not hasattr(self, "review_filter_combo"):
            return True
        mode = self.review_filter_combo.currentText()
        return mode in {"All", "Unreviewed"}

    def _review_reset_random_state(self: "MainController") -> None:
        self._review_random_seen_indices.clear()
        self._review_random_history.clear()

    def _review_pick_random_item_index(self: "MainController") -> int | None:
        pending = [
            idx
            for idx in self._review_filtered_indices
            if not bool(self._review_items[idx].get("completed"))
        ]
        if not pending:
            return None
        candidates = [idx for idx in pending if idx not in self._review_random_seen_indices]
        if not candidates:
            self._review_random_seen_indices.clear()
            candidates = pending[:]
        current = self._review_current_item_index
        if current in candidates and len(candidates) > 1:
            candidates = [idx for idx in candidates if idx != current]
        if not candidates:
            return None
        return self._review_rng.choice(candidates)

    def _review_load_filtered_pos(
        self: "MainController",
        pos: int,
        force: bool = False,
        *,
        record_history: bool = True,
    ) -> None:
        if not self._review_filtered_indices:
            return
        if pos < 0 or pos >= len(self._review_filtered_indices):
            return
        item_index = self._review_filtered_indices[pos]
        if not force and item_index == self._review_current_item_index:
            return
        if not self._prompt_save_if_dirty():
            return

        item = self._review_items[item_index]
        try:
            load_mask_path = item["corrected_path"] or item["pred_path"]
            target_shape = self._review_stack_shape(load_mask_path)
            self.model.load(item["raw_path"])
            raw_resample_note = ""
            if target_shape and self.model.data is not None:
                src_shape = tuple(int(v) for v in self.model.data.shape)
                if src_shape != target_shape:
                    self.model.resample_image_to_shape(target_shape)
                    raw_resample_note = (
                        f"Raw aligned {src_shape[0]}->{target_shape[0]} "
                        "(mask-grid review)."
                    )
            self.model.load_masks(load_mask_path)
        except Exception as exc:
            QMessageBox.warning(self, "Review Load", f"Failed to load review pair:\n{exc}")
            return

        # Default save target: existing corrected path or a deterministic path.
        self.model.mask_path = item["corrected_path"] or self._review_default_corrected_path(item)
        self.model.mask_dirty = False
        self.slider.setRange(0, self.model.n_slices - 1)
        self.slider.setEnabled(True)
        self._review_current_filtered_pos = pos
        self._review_current_item_index = item_index
        if self._review_should_random_select():
            self._review_random_seen_indices.add(item_index)
            if record_history:
                if not self._review_random_history or self._review_random_history[-1] != item_index:
                    self._review_random_history.append(item_index)
        self._sync_review_grade_from_item(item)
        self._update_view(reset_view=True)
        self._review_update_info_label()
        notes = []
        if raw_resample_note:
            notes.append(raw_resample_note)
        if self.model.mask_alignment_note:
            notes.append(self.model.mask_alignment_note)
        done_mark = "DONE" if item.get("completed") else "PENDING"
        self.statusBar().showMessage(
            f"Loaded review item {pos + 1}/{len(self._review_filtered_indices)}: {item['zstack_id']} [{done_mark}]"
            + (f" | {' '.join(notes)}" if notes else "")
        )

    # --------- ui helpers ---------
    def _set_review_controls_enabled(self: "MainController", enabled: bool) -> None:
        widgets = [
            "review_prev_stack_btn",
            "review_next_stack_btn",
            "review_filter_combo",
            "review_grade_combo",
            "review_mark_a_btn",
            "review_mark_b_btn",
            "review_mark_c_btn",
            "review_save_corrected_btn",
            "review_export_final_btn",
        ]
        for name in widgets:
            widget = getattr(self, name, None)
            if widget is not None:
                widget.setEnabled(enabled)

    def _sync_review_grade_from_item(self: "MainController", item: dict) -> None:
        if not hasattr(self, "review_grade_combo"):
            return
        grade = normalize_review_grade(item.get("grade"))
        text = grade if grade else "Unreviewed"
        self._review_grade_updating = True
        self.review_grade_combo.setCurrentText(text)
        self._review_grade_updating = False

    def _review_sync_grade_combo(self: "MainController") -> None:
        item = self._review_current_item()
        if item is None:
            return
        self._sync_review_grade_from_item(item)

    def _review_update_info_label(self: "MainController") -> None:
        if not hasattr(self, "review_info_label"):
            return
        total = len(self._review_items)
        filtered = len(self._review_filtered_indices)
        if total == 0:
            self.review_info_label.setText("Review: tracker not loaded")
            return
        counts = {"A": 0, "B": 0, "C": 0, "U": 0, "DONE": 0}
        for item in self._review_items:
            grade = normalize_review_grade(item.get("grade"))
            if grade in {"A", "B", "C"}:
                counts[grade] += 1
            else:
                counts["U"] += 1
            if item.get("completed"):
                counts["DONE"] += 1

        if self._review_current_item_index < 0:
            self.review_info_label.setText(
                (
                    f"Review: 0/{filtered} (total {total}) | "
                    f"A:{counts['A']} B:{counts['B']} C:{counts['C']} U:{counts['U']} "
                    f"Done:{counts['DONE']}"
                )
            )
            return

        current_item = self._review_items[self._review_current_item_index]
        grade = normalize_review_grade(current_item.get("grade")) or "U"
        done_tag = "DONE" if current_item.get("completed") else "PENDING"
        self.review_info_label.setText(
            (
                f"Review: {self._review_current_filtered_pos + 1}/{filtered} (total {total}) | "
                f"{current_item['zstack_id']} [{grade}/{done_tag}] | "
                f"A:{counts['A']} B:{counts['B']} C:{counts['C']} U:{counts['U']} Done:{counts['DONE']}"
            )
        )

    def _review_pick_initial_filter(self: "MainController") -> str:
        has_pending_unreviewed = any(
            (not normalize_review_grade(item.get("grade"))) and (not bool(item.get("completed")))
            for item in self._review_items
        )
        return "Unreviewed" if has_pending_unreviewed else "All"

    def _review_set_filter(self: "MainController", value: str) -> None:
        if not hasattr(self, "review_filter_combo"):
            return
        self._review_filter_updating = True
        self.review_filter_combo.setCurrentText(value)
        self._review_filter_updating = False

    # --------- table helpers ---------
    def _review_rebuild_header_index(self: "MainController") -> None:
        self._review_headers = {name: idx + 1 for idx, name in enumerate(self._review_header_names)}

    def _review_normalize_table_rows(self: "MainController") -> None:
        for row in self._review_table_rows:
            for name in self._review_header_names:
                row.setdefault(name, "")

    def _review_row_numbers(self: "MainController") -> range:
        return range(2, len(self._review_table_rows) + 2)

    def _review_cell(self: "MainController", row: int, column_name: str) -> str:
        if column_name not in self._review_headers:
            return ""
        idx = row - 2
        if idx < 0 or idx >= len(self._review_table_rows):
            return ""
        value = self._review_table_rows[idx].get(column_name, "")
        if value is None:
            return ""
        return str(value).strip()

    def _review_set_cell(self: "MainController", row: int, column_name: str, value: object) -> None:
        if column_name not in self._review_headers:
            return
        idx = row - 2
        if idx < 0:
            return
        while idx >= len(self._review_table_rows):
            self._review_table_rows.append({name: "" for name in self._review_header_names})
        self._review_table_rows[idx][column_name] = value

    def _review_first_nonempty(self: "MainController", row: int, columns: list[str]) -> str:
        for name in columns:
            value = self._review_cell(row, name)
            if value:
                return value
        return ""

    def _review_default_corrected_path(self: "MainController", item: dict) -> str:
        tracker_dir = os.path.dirname(self._review_tracker_path or "")
        grade = normalize_review_grade(item.get("grade"))
        grade_folder = grade if grade else "UNREVIEWED"
        out_dir = os.path.join(tracker_dir, "review_corrected_masks", grade_folder)
        os.makedirs(out_dir, exist_ok=True)
        zstack_id = item.get("zstack_id") or f"row{item['row']}"
        return os.path.join(out_dir, f"{zstack_id}_review_mask.tif")

    def _review_build_corrected_mask_metadata(
        self: "MainController", item: dict, out_path: str
    ) -> dict:
        """Build pairing metadata to embed in corrected mask TIFF."""
        raw_path = item.get("raw_path") or ""
        pred_path = item.get("pred_path") or ""
        loaded_mask_path = item.get("corrected_path") or pred_path

        raw_shape = self._review_stack_shape(raw_path) if raw_path else None
        review_shape = tuple(int(v) for v in self.model.data.shape) if self.model.data is not None else None
        mask_shape = tuple(int(v) for v in self.model.masks.shape) if self.model.masks is not None else None
        pixel_sizes = self.model.get_pixel_sizes()

        scale_zyx = None
        if raw_shape and mask_shape and raw_shape[0] > 0 and raw_shape[1] > 0 and raw_shape[2] > 0:
            scale_zyx = [
                float(mask_shape[0]) / float(raw_shape[0]),
                float(mask_shape[1]) / float(raw_shape[1]),
                float(mask_shape[2]) / float(raw_shape[2]),
            ]

        return {
            "schema": "myelin_review_pairing_v1",
            "saved_at": self._review_now(),
            "tool": "Myelin_anno_tool",
            "review": {
                "zstack_id": item.get("zstack_id"),
                "grade": normalize_review_grade(item.get("grade")) or "UNREVIEWED",
                "completed": bool(item.get("completed")),
                "row": item.get("row"),
                "tracker_path_windows": local_to_windows_path(self._review_tracker_path or ""),
            },
            "pairing": {
                "source_raw_id": item.get("zstack_id"),
                "raw_path_local": raw_path,
                "raw_path_windows": local_to_windows_path(raw_path),
                "prediction_path_local": pred_path,
                "prediction_path_windows": local_to_windows_path(pred_path),
                "prediction_file_name": os.path.basename(pred_path) if pred_path else "",
                "loaded_mask_source_local": loaded_mask_path,
                "loaded_mask_source_windows": local_to_windows_path(loaded_mask_path),
                "corrected_mask_output_local": out_path,
                "corrected_mask_output_windows": local_to_windows_path(out_path),
            },
            "dimensions": {
                "raw_original_zyx": list(raw_shape) if raw_shape else None,
                "review_image_zyx": list(review_shape) if review_shape else None,
                "saved_mask_zyx": list(mask_shape) if mask_shape else None,
                "raw_resampled_for_review": bool(raw_shape and review_shape and raw_shape != review_shape),
                "resample_policy": "raw_to_mask_grid",
                "raw_to_mask_scale_zyx": scale_zyx,
            },
            "physical_size_um": {
                "review_grid_xyz": list(pixel_sizes) if pixel_sizes else None,
            },
        }

    @staticmethod
    def _review_stack_shape(path: str) -> tuple[int, int, int] | None:
        """Return stack shape as (Z, Y, X), squeezing singleton dims when needed."""
        try:
            with tifffile.TiffFile(path) as tf:
                if tf.series:
                    shape = tuple(int(v) for v in tf.series[0].shape)
                else:
                    page = tf.pages[0]
                    shape = (len(tf.pages), int(page.shape[0]), int(page.shape[1]))
        except Exception:
            return None

        if len(shape) == 2:
            return (1, shape[0], shape[1])
        if len(shape) == 3:
            return shape
        if len(shape) == 4:
            # Common cases: (1, Z, Y, X) or (Z, 1, Y, X)
            squeezed = tuple(v for v in shape if v != 1)
            if len(squeezed) == 3:
                return squeezed
            return shape[-3:]
        return None

    def _review_save_tracker(self: "MainController") -> bool:
        if not self._review_tracker_path:
            return False
        try:
            self._review_write_tracker(
                self._review_tracker_path,
                self._review_header_names,
                self._review_table_rows,
            )
            return True
        except PermissionError:
            QMessageBox.warning(
                self,
                "Review Tracker",
                "Cannot write tracker file (it may be open in another app). Close it and retry.",
            )
        except Exception as exc:
            QMessageBox.warning(self, "Review Tracker", f"Failed to save tracker:\n{exc}")
        return False

    def _review_current_item(self: "MainController") -> dict | None:
        if self._review_current_item_index < 0:
            return None
        if self._review_current_item_index >= len(self._review_items):
            return None
        return self._review_items[self._review_current_item_index]

    @staticmethod
    def _review_now() -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
